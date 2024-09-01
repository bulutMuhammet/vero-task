import argparse
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from datetime import datetime


def fetch_data(file_path):
    url = "http://localhost:5000/upload"
    with open(file_path, 'rb') as file:
        response = requests.post(url, files={'file': file})

    return response.json()


def create_excel(data, keys, colored):
    wb = Workbook()
    ws = wb.active

    headers = ['rnr', 'gruppe'] + keys
    ws.append(headers)

    header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.border = border

    color_map = {
        'green': '007500',
        'orange': 'FFA500',
        'red': 'b30000'
    }

    for item in data:
        row = [item.get('rnr'), item.get('gruppe')] + [item.get(key, '') for key in keys]
        ws.append(row)

        if colored:
            hu_date = pd.to_datetime(item.get('hu'))
            months_diff = (datetime.now() - hu_date).days // 30

            if months_diff <= 3:
                fill_color = PatternFill(start_color=color_map['green'], end_color=color_map['green'],
                                         fill_type="solid")
            elif months_diff <= 12:
                fill_color = PatternFill(start_color=color_map['orange'], end_color=color_map['orange'],
                                         fill_type="solid")
            else:
                fill_color = PatternFill(start_color=color_map['red'], end_color=color_map['red'], fill_type="solid")

            for cell in ws[ws.max_row]:
                cell.fill = fill_color

        if 'labelColors' in item:
            for col_index, color in enumerate(item['labelColors'], start=len(headers) + 1):
                cell = ws.cell(row=ws.max_row, column=col_index)
                cell.fill = PatternFill(start_color=color.replace('#', ''), end_color=color.replace('#', ''),
                                        fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    file_name = f'vehicles_{datetime.now().date()}.xlsx'
    wb.save(file_name)
    print(f"Excel file saved as '{file_name}'")


def main():
    parser = argparse.ArgumentParser(description='Client for vehicle data API.')
    parser.add_argument('-f', '--file', required=True, help='Path to the CSV file')
    parser.add_argument('-k', '--keys', nargs='*', default=[], help='Keys to include in additional columns')
    parser.add_argument('-c', '--colored', type=bool, default=True, help='Whether to color rows based on age')

    args = parser.parse_args()

    data = fetch_data(args.file, args.keys, args.colored)
    create_excel(data, args.keys, args.colored)


if __name__ == "__main__":
    main()
